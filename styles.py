"""Shared style presets for Veriscope Edge — 4-color palette."""
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Protection

COLORS = {
    'bg_dark': '0B1120',
    'surface': '111827',
    'accent_gold': 'D4AF37',
    'profit_green': '22D3A6',
    'loss_red': 'E5484D',
    'text_light': 'F8FAFC',
    'text_dark': '1E293B',
    'input_bg': 'DCEBFF',
}

THIN = Side(style='thin', color=COLORS['bg_dark'])
BORDER_ALL = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


class StylePresets:
    @staticmethod
    def title_band():
        return {
            'font': Font(name='Calibri', size=16, bold=True, color=COLORS['accent_gold']),
            'fill': PatternFill('solid', fgColor=COLORS['bg_dark']),
            'alignment': Alignment(horizontal='center', vertical='center'),
        }

    @staticmethod
    def section_header():
        return {
            'font': Font(size=12, bold=True, color=COLORS['text_light']),
            'fill': PatternFill('solid', fgColor=COLORS['surface']),
            'alignment': Alignment(horizontal='left', vertical='center'),
        }

    @staticmethod
    def table_header():
        return {
            'font': Font(size=10, bold=True, color=COLORS['text_light']),
            'fill': PatternFill('solid', fgColor=COLORS['bg_dark']),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': BORDER_ALL,
        }

    @staticmethod
    def input_cell():
        return {
            'font': Font(size=10, color=COLORS['text_dark']),
            'fill': PatternFill('solid', fgColor=COLORS['input_bg']),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': BORDER_ALL,
            'protection': Protection(locked=False),
        }

    @staticmethod
    def locked_cell():
        return {
            'font': Font(size=10, color=COLORS['text_dark']),
            'fill': PatternFill('solid', fgColor='FFFFFF'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': BORDER_ALL,
            'protection': Protection(locked=True),
        }

    @staticmethod
    def kpi_label():
        return {
            'font': Font(size=9, bold=True, color=COLORS['text_light']),
            'fill': PatternFill('solid', fgColor=COLORS['surface']),
            'alignment': Alignment(horizontal='center', vertical='center'),
        }

    @staticmethod
    def kpi_value(color=COLORS['accent_gold']):
        return {
            'font': Font(size=20, bold=True, color=color),
            'fill': PatternFill('solid', fgColor=COLORS['bg_dark']),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(top=Side(style='medium', color=color), bottom=Side(style='medium', color=color),
                              left=Side(style='medium', color=color), right=Side(style='medium', color=color)),
        }

    @staticmethod
    def warning():
        return {
            'font': Font(size=9, italic=True, color=COLORS['bg_dark']),
            'fill': PatternFill('solid', fgColor='FDE68A'),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
        }


def apply_style(cell, style: dict):
    if 'font' in style:
        cell.font = style['font']
    if 'fill' in style:
        cell.fill = style['fill']
    if 'alignment' in style:
        cell.alignment = style['alignment']
    if 'border' in style:
        cell.border = style['border']
    if 'protection' in style:
        cell.protection = style['protection']
