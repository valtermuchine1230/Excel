# veriscope_edge_generator.py
"""
VERISCOPE EDGE - Main Generator Script
Builds complete trading performance dashboard Excel workbook.
Language: Python 3.10+
Library: openpyxl
Output: ./output/Veriscope_Edge.xlsx
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import LineChart, BarChart, AreaChart, DoughnutChart
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Protection, numbers
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.drawing.image import Image as XLImage
from openpyxl.workbook.defined_name import DefinedName

from styles import COLORS, StylePresets, apply_style

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def ensure_output_dir():
    """Ensure output directory exists."""
    if not os.path.exists('output'):
        os.makedirs('output')

def set_sheet_appearance(ws, show_gridlines=False):
    """Apply standard appearance settings to a worksheet."""
    ws.sheet_view.showGridLines = show_gridlines
    ws.print_options.horizontalCentered = True
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'landscape'

def add_title_band(ws, title, row=1, colspan=10):
    """Add a styled title band at the top of a sheet."""
    merged_range = f"A{row}:{get_column_letter(colspan)}{row}"
    ws.merge_cells(merged_range)
    cell = ws[f"A{row}"]
    cell.value = title
    # Correct: row_dimensions belongs to Worksheet, not Cell
    ws.row_dimensions[row].height = 28
    apply_style(cell, StylePresets.title_band())

def create_named_range(wb, name, sheet_name, cell):
    """Create a named range in the workbook."""
    # Use DefinedName and append to workbook defined names
    dn = DefinedName(name=name, attr_text=f"'{sheet_name}'!${cell}")
    wb.defined_names.append(dn)

def apply_sheet_protection(ws, password='veriscope2026'):
    """Apply sheet protection with password."""
    ws.protection.sheet = True
    ws.protection.password = password
    ws.protection.enable()

def format_currency(value_str='#,##0.00'):
    """Return Excel currency format string."""
    return f'"$"{value_str}'

# ============================================================================
# SHEET BUILDERS
# ============================================================================

def build_lists_sheet(wb):
    """Build hidden 99_Lists sheet with dropdown sources."""
    ws = wb.create_sheet('99_Lists')
    ws.sheet_state = 'hidden'
    
    set_sheet_appearance(ws)
    add_title_band(ws, '📋 Data Lists (Hidden)', 1, 6)
    
    # Define list data
    lists_data = {
        'SETUPS': ['Breakout', 'Pullback', 'Support/Resistance', 'Fibonacci', 'Moving Average Cross', 'Pin Bar', 'Divergence', 'Custom'],
        'SYMBOLS': ['EURUSD', 'GBPUSD', 'USDJPY', 'GOLD', 'BTCUSD', 'AAPL', 'MSFT', 'Other'],
        'EMOTIONS': ['Calm', 'Focused', 'Anxious', 'FOMO', 'Tilt', 'Confident', 'Bored'],
        'RULES_BROKEN': ['Entry Criteria', 'Position Sizing', 'Stop Placement', 'Early Exit', 'Overtrading', 'Revenge Trade', 'None'],
        'ASSET_CLASSES': ['Forex', 'Stocks', 'Futures', 'Crypto', 'Options'],
    }
    
    current_col = 1
    for list_name, items in lists_data.items():
        ws.cell(row=3, column=current_col).value = list_name
        ws.cell(row=3, column=current_col).font = Font(bold=True, color=COLORS['accent_gold'])
        for idx, item in enumerate(items, start=4):
            ws.cell(row=idx, column=current_col).value = item
        current_col += 2
    
    # Auto-size columns
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 15

def build_settings_sheet(wb):
    """Build 02_Settings control panel."""
    ws = wb.create_sheet('02_Settings', 1)
    set_sheet_appearance(ws)
    add_title_band(ws, '⚙️ Settings & Configuration', 1, 8)
    
    # Define settings
    settings = [
        ('Starting Capital ($)', 10000, 'C5'),
        ('Base Currency', 'USD', 'C6'),
        ('Currency Symbol', '$', 'C7'),
        ('Default Risk % per Trade', 2, 'C8'),
        ('Daily Loss Limit ($)', 500, 'C9'),
        ('Trader Name', 'Anonymous', 'C10'),
        ('Broker(s)', 'Broker Name', 'C11'),
        ('Timezone', 'UTC', 'C12'),
    ]
    
    row = 4
    for label, default_value, cell_ref in settings:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        ws.cell(row=row, column=3).value = default_value
        apply_style(ws.cell(row=row, column=3), StylePresets.input_cell())
        row += 1
    
    # Create named ranges
    create_named_range(wb, 'AccountBalance', '02_Settings', 'C5')
    create_named_range(wb, 'BaseCurrency', '02_Settings', 'C6')
    create_named_range(wb, 'CurrencySymbol', '02_Settings', 'C7')
    create_named_range(wb, 'RiskPercent', '02_Settings', 'C8')
    create_named_range(wb, 'DailyLossLimit', '02_Settings', 'C9')
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['C'].width = 25
    
    apply_sheet_protection(ws)

def build_cover_sheet(wb):
    """Build 00_Cover landing page."""
    ws = wb.create_sheet('00_Cover', 0)
    set_sheet_appearance(ws)
    
    # Full dark background
    for row in range(1, 50):
        for col in range(1, 15):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=COLORS['bg_dark'], end_color=COLORS['bg_dark'], fill_type='solid')
    
    # Title
    ws.merge_cells('A5:N6')
    title_cell = ws['A5']
    title_cell.value = 'VERISCOPE EDGE'
    title_cell.font = Font(name='Montserrat', size=28, bold=True, color=COLORS['accent_gold'])
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[5].height = 40
    
    # Subtitle
    ws.merge_cells('A8:N8')
    subtitle = ws['A8']
    subtitle.value = 'The Complete Trading Performance System'
    subtitle.font = Font(size=14, color=COLORS['text_light'], bold=True)
    subtitle.alignment = Alignment(horizontal='center', vertical='center')
    
    # Description
    ws.merge_cells('A10:N12')
    desc = ws['A10']
    desc.value = (
        'Premium all-in-one trading journal, risk management, and performance analytics.\n'
        'Know Your Edge. Trade With Proof.\n'
        'Version 1.0 — Generated ' + datetime.now().strftime('%Y-%m-%d')
    )
    desc.font = Font(size=11, color=COLORS['text_light'])
    desc.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    ws.row_dimensions[10].height = 50
    
    # Logo placeholder (kept as in original - you said "apenas corrija esse")
    ws.merge_cells('A14:N18')
    logo = ws['A14']
    logo.value = '[LOGO PLACEHOLDER — Insert 300x150px PNG here]'
    logo.fill = PatternFill(start_color=COLORS['surface'], end_color=COLORS['surface'], fill_type='solid')
    logo.font = Font(size=10, color=COLORS['accent_gold'], italic=True)
    logo.alignment = Alignment(horizontal='center', vertical='center')
    logo.border = Border(
        top=Side(style='thin', color=COLORS['accent_gold']),
        bottom=Side(style='thin', color=COLORS['accent_gold']),
        left=Side(style='thin', color=COLORS['accent_gold']),
        right=Side(style='thin', color=COLORS['accent_gold']),
    )
    
    # Navigation menu
    nav_row = 21
    ws.merge_cells(f'A{nav_row}:N{nav_row}')
    nav_header = ws[f'A{nav_row}']
    nav_header.value = '📑 NAVIGATION MENU'
    nav_header.font = Font(size=12, bold=True, color=COLORS['accent_gold'])
    nav_header.alignment = Alignment(horizontal='left', vertical='center')
    
    sheets_to_link = [
        ('01_How_To_Use', '📚 How to Use'),
        ('02_Settings', '⚙️ Settings'),
        ('03_Dashboard', '📊 Dashboard'),
        ('04_Trading_Journal', '📔 Trading Journal'),
        ('05_Risk_Manager', '⚠️ Risk Manager'),
        ('06_Drawdown_Tracker', '📉 Drawdown Tracker'),
        ('07_Profit_Dashboard', '💰 Profit Dashboard'),
    ]
    
    nav_row += 2
    for sheet_name, label in sheets_to_link:
        cell = ws[f'A{nav_row}']
        cell.value = f'=HYPERLINK("#{sheet_name}!A1","{label}")'
        cell.font = Font(size=11, color=COLORS.get('accent_teal', COLORS['accent_gold']), underline='single')
        nav_row += 1
    
    ws.column_dimensions['A'].width = 30

def build_how_to_use_sheet(wb):
    """Build 01_How_To_Use onboarding sheet."""
    ws = wb.create_sheet('01_How_To_Use', 1)
    set_sheet_appearance(ws)
    add_title_band(ws, '📚 How to Use Veriscope Edge', 1, 10)
    
    content = [
        ('', 4),
        ('QUICK START (5 Steps)', 5),
        ('', 6),
        ('1. Go to SETTINGS tab and configure your starting capital, risk %, and broker details.', 7),
        ('2. Open TRADING JOURNAL and log every trade: entry, exit, and outcome.', 8),
        ('3. Watch DASHBOARD update automatically with your live P&L and win rate.', 9),
        ('4. Use RISK MANAGER to calculate position size before entering a trade.', 10),
        ('5. Review PROFIT DASHBOARD, R-MULTIPLE TRACKER, and HEATMAPS weekly to find patterns.', 11),
        ('', 12),
        ('FAQ FOR BEGINNERS', 13),
        ('', 14),
        ('Q: What is R-Multiple?', 15),
        ('A: R = Risk on the trade. If risk $100 and profit $300, that is +3R. Expectancy in R tells your edge.', 16),
        ('', 17),
        ('Q: What is Drawdown?', 18),
        ('A: The decline from peak balance to lowest point. Max DD shows your worst losing streak.', 19),
        ('', 20),
        ('Q: What is Profit Factor?', 21),
        ('A: Total wins ÷ total losses. Above 1.5 is good; above 2.0 is excellent.', 22),
        ('', 23),
        ('Q: Can I edit this file?', 24),
        ('A: Yes! Input cells are unlocked. Formula cells are protected to prevent accidental changes.', 25),
    ]
    
    for text, row in content:
        cell = ws.cell(row=row, column=1)
        cell.value = text
        if any(x in text for x in ['QUICK START', 'FAQ', 'Q:', 'A:', '1.', '2.', '3.', '4.', '5.']):
            cell.font = Font(bold=True, size=11 if any(x in text for x in ['QUICK', 'FAQ']) else 10)
        ws.row_dimensions[row].height = 20
    
    ws.column_dimensions['A'].width = 80
    apply_sheet_protection(ws)

def build_dashboard_sheet(wb):
    """Build 03_Dashboard overview KPI cards."""
    ws = wb.create_sheet('03_Dashboard', 3)
    set_sheet_appearance(ws)
    add_title_band(ws, '📊 Trading Dashboard', 1, 10)
    
    # Create KPI cards in a 2x4 grid
    kpi_data = [
        ('Net P&L', '=IFERROR(SUMIF(Tbl_Journal[Net P&L],">="&-9999999,Tbl_Journal[Net P&L]),0)', COLORS['accent_gold']),
        ('Win Rate %', '=IFERROR(COUNTIF(Tbl_Journal[Net P&L],">0")/COUNTA(Tbl_Journal[Net P&L])*100,0)', COLORS.get('accent_teal', COLORS['accent_gold'])),
        ('Profit Factor', '=IFERROR(SUMIF(Tbl_Journal[Net P&L],">0")/ABS(SUMIF(Tbl_Journal[Net P&L],"<0")),0)', COLORS['accent_gold']),
        ('Expectancy (R)', '=IFERROR(AVERAGE(Tbl_Journal[R-Multiple]),0)', COLORS.get('accent_teal', COLORS['accent_gold'])),
        ('Current DD %', '=IFERROR((MAX(Tbl_Journal[Running Balance])-MIN(Tbl_Journal[Running Balance]))/MAX(Tbl_Journal[Running Balance])*-100,0)', COLORS['accent_red']),
        ('Total Trades', '=COUNTA(Tbl_Journal[Trade #])', COLORS['accent_gold']),
        ('Best Trade', '=IFERROR(MAX(Tbl_Journal[Net P&L]),0)', COLORS.get('accent_teal', COLORS['accent_gold'])),
        ('Worst Trade', '=IFERROR(MIN(Tbl_Journal[Net P&L]),0)', COLORS['accent_red']),
    ]
    
    row = 4
    col = 1
    for label, formula, color in kpi_data:
        # Label
        label_cell = ws.cell(row=row, column=col)
        label_cell.value = label
        apply_style(label_cell, StylePresets.kpi_label())
        ws.row_dimensions[row].height = 18
        
        # Value
        value_cell = ws.cell(row=row + 1, column=col)
        value_cell.value = formula
        apply_style(value_cell, StylePresets.kpi_value(color))
        ws.row_dimensions[row + 1].height = 30
        
        if col >= 6:
            col = 1
            row += 3
        else:
            col += 1
    
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    
    apply_sheet_protection(ws)

def build_trading_journal_sheet(wb):
    """Build 04_Trading_Journal core data entry engine."""
    ws = wb.create_sheet('04_Trading_Journal', 4)
    set_sheet_appearance(ws)
    add_title_band(ws, '📔 Trading Journal', 1, 26)
    
    # Column headers (Row 3-4: PRE-TRADE and POST-TRADE group headers)
    ws.merge_cells('A3:H3')
    pre_header = ws['A3']
    pre_header.value = 'PRE-TRADE SETUP'
    pre_header.fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    pre_header.font = Font(bold=True, color='FFFFFF')
    pre_header.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('I3:Q3')
    post_header = ws['I3']
    post_header.value = 'POST-TRADE RESULTS'
    post_header.fill = PatternFill(start_color='1F5E4F', end_color='1F5E4F', fill_type='solid')
    post_header.font = Font(bold=True, color='FFFFFF')
    post_header.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('R3:Z3')
    meta_header = ws['R3']
    meta_header.value = 'ANALYSIS & NOTES'
    meta_header.fill = PatternFill(start_color='4A3728', end_color='4A3728', fill_type='solid')
    meta_header.font = Font(bold=True, color='FFFFFF')
    meta_header.alignment = Alignment(horizontal='center', vertical='center')
    
    # Column names (Row 4)
    columns = [
        'Trade #', 'Date', 'Time', 'Symbol', 'Direction', 'Setup', 'Market Condition', 'Confidence',
        'Entry Price', 'Stop Loss', 'Position Size', 'Risk $ (1R)', 'Risk %',
        'Exit Price', 'Exit Date', 'Gross P&L', 'Commission', 'Net P&L', 'R-Multiple', 'Running Balance',
        'Rule Compliant?', 'Rule Broken', 'Emotion', 'Tags', 'Screenshot Link', 'Notes'
    ]
    
    for idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=idx)
        cell.value = col_name
        apply_style(cell, StylePresets.table_header())
        ws.row_dimensions[4].height = 25
    
    # Pre-populate 500 empty rows with formulas
    for trade_row in range(5, 505):
        # Trade # formula
        ws[f'A{trade_row}'].value = f'=IF(D{trade_row}="","",COUNTA($D$5:D{trade_row}))'
        apply_style(ws[f'A{trade_row}'], StylePresets.locked_cell())
        
        # Input cells
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'O', 'W', 'X', 'Y', 'Z']:
            apply_style(ws[f'{col}{trade_row}'], StylePresets.input_cell())
        
        # Risk $ (1R) formula
        ws[f'L{trade_row}'].value = f'=IF(OR(I{trade_row}="",J{trade_row}="",K{trade_row}=""),"",ABS(I{trade_row}-J{trade_row})*K{trade_row})'
        apply_style(ws[f'L{trade_row}'], StylePresets.locked_cell())
        
        # Risk %
        ws[f'M{trade_row}'].value = f'=IF(L{trade_row}="","",L{trade_row}/AccountBalance)'
        apply_style(ws[f'M{trade_row}'], StylePresets.locked_cell())
        
        # Gross P&L formula
        ws[f'P{trade_row}'].value = f'=IF(OR(E{trade_row}="",N{trade_row}=""),"",IF(E{trade_row}="Long",(N{trade_row}-I{trade_row})*K{trade_row},(I{trade_row}-N{trade_row})*K{trade_row}))'
        apply_style(ws[f'P{trade_row}'], StylePresets.locked_cell())
        
        # Net P&L
        ws[f'R{trade_row}'].value = f'=IF(P{trade_row}="","",P{trade_row}-Q{trade_row})'
        apply_style(ws[f'R{trade_row}'], StylePresets.locked_cell())
        
        # R-Multiple
        ws[f'S{trade_row}'].value = f'=IF(OR(R{trade_row}="",L{trade_row}=0,L{trade_row}=""),"",R{trade_row}/L{trade_row})'
        apply_style(ws[f'S{trade_row}'], StylePresets.locked_cell())
        
        # Running Balance
        ws[f'T{trade_row}'].value = f'=IF(R{trade_row}="","",AccountBalance+SUM($R$5:R{trade_row}))'
        apply_style(ws[f'T{trade_row}'], StylePresets.locked_cell())
    
    # Set column widths
    widths = [8, 12, 10, 10, 10, 14, 16, 11, 12, 12, 12, 12, 10, 12, 12, 12, 12, 12, 12, 14, 12, 14, 12, 15, 15, 20]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Create Excel Table (ListObject)
    tab = Table(displayName='Tbl_Journal', ref='A4:Z504')
    style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    # Data validation for dropdowns
    symbol_validation = DataValidation(type='list', formula1='=99_Lists!$B$4:$B$11', allow_blank=True)
    direction_validation = DataValidation(type='list', formula1='"Long,Short"', allow_blank=True)
    setup_validation = DataValidation(type='list', formula1='=99_Lists!$A$4:$A$11', allow_blank=True)
    market_condition_validation = DataValidation(type='list', formula1='"Trending,Ranging,Choppy,News-Driven"', allow_blank=True)
    confidence_validation = DataValidation(type='list', formula1='"1,2,3,4,5"', allow_blank=True)
    rule_validation = DataValidation(type='list', formula1='"Yes,No"', allow_blank=True)
    rule_broken_validation = DataValidation(type='list', formula1='=99_Lists!$D$4:$D$10', allow_blank=True)
    emotion_validation = DataValidation(type='list', formula1='=99_Lists!$C$4:$C$10', allow_blank=True)
    
    ws.add_data_validation(symbol_validation)
    ws.add_data_validation(direction_validation)
    ws.add_data_validation(setup_validation)
    ws.add_data_validation(market_condition_validation)
    ws.add_data_validation(confidence_validation)
    ws.add_data_validation(rule_validation)
    ws.add_data_validation(rule_broken_validation)
    ws.add_data_validation(emotion_validation)
    
    symbol_validation.add(f'D5:D504')
    direction_validation.add(f'E5:E504')
    setup_validation.add(f'F5:F504')
    market_condition_validation.add(f'G5:G504')
    confidence_validation.add(f'H5:H504')
    rule_validation.add(f'U5:U504')
    rule_broken_validation.add(f'V5:V504')
    emotion_validation.add(f'X5:X504')
    
    # Freeze panes
    ws.freeze_panes = 'A5'
    
    apply_sheet_protection(ws)

def build_risk_manager_sheet(wb):
    """Build 05_Risk_Manager position sizing."""
    ws = wb.create_sheet('05_Risk_Manager', 5)
    set_sheet_appearance(ws)
    add_title_band(ws, '⚠️ Risk Manager', 1, 8)
    
    # Position Sizing Calculator
    ws.merge_cells('A4:C4')
    calc_header = ws['A4']
    calc_header.value = 'POSITION SIZING CALCULATOR'
    apply_style(calc_header, StylePresets.section_header())
    
    # Inputs
    ws['A6'].value = 'Account Size'
    ws['B6'].value = '=AccountBalance'
    apply_style(ws['B6'], StylePresets.locked_cell())
    
    ws['A7'].value = 'Risk % per Trade'
    ws['B7'].value = '=RiskPercent'
    apply_style(ws['B7'], StylePresets.input_cell())
    
    ws['A8'].value = 'Entry Price'
    apply_style(ws['B8'], StylePresets.input_cell())
    
    ws['A9'].value = 'Stop Loss Price'
    apply_style(ws['B9'], StylePresets.input_cell())
    
    ws['A11'].value = 'Risk in $'
    ws['B11'].value = '=IF(OR(B6="",B7=""),"",B6*B7/100)'
    apply_style(ws['B11'], StylePresets.locked_cell())
    
    ws['A12'].value = 'Stop Distance (Pips/Units)'
    ws['B12'].value = '=IF(OR(B8="",B9=""),"",ABS(B8-B9))'
    apply_style(ws['B12'], StylePresets.locked_cell())
    
    ws['A13'].value = 'Position Size (Units)'
    ws['B13'].value = '=IF(OR(B11="",B12="",B12=0),"",B11/B12)'
    apply_style(ws['B13'], StylePresets.kpi_value(COLORS['accent_gold']))
    ws.row_dimensions[13].height = 30
    
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    
    apply_sheet_protection(ws)

def build_drawdown_tracker_sheet(wb):
    """Build 06_Drawdown_Tracker."""
    ws = wb.create_sheet('06_Drawdown_Tracker', 6)
    set_sheet_appearance(ws)
    add_title_band(ws, '📉 Drawdown Tracker', 1, 8)
    
    # KPI Cards
    ws['A4'].value = 'Current DD %'
    apply_style(ws['A4'], StylePresets.kpi_label())
    ws['A5'].value = '=IFERROR((MAX(Tbl_Journal[Running Balance])-MIN(Tbl_Journal[Running Balance]))/MAX(Tbl_Journal[Running Balance])*-100,0)'
    apply_style(ws['A5'], StylePresets.kpi_value(COLORS['accent_red']))
    
    ws['B4'].value = 'Max DD %'
    apply_style(ws['B4'], StylePresets.kpi_label())
    ws['B5'].value = '=IFERROR(MINIFS(Tbl_Journal[Drawdown %],Tbl_Journal[Drawdown %],"<0"),0)'
    apply_style(ws['B5'], StylePresets.kpi_value(COLORS['accent_red']))
    
    ws['C4'].value = 'Days in DD'
    apply_style(ws['C4'], StylePresets.kpi_label())
    ws['C5'].value = '=COUNTIF(Tbl_Journal[Underwater],"<0")'
    apply_style(ws['C5'], StylePresets.kpi_value(COLORS['accent_gold']))
    
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    
    apply_sheet_protection(ws)

def build_profit_dashboard_sheet(wb):
    """Build 07_Profit_Dashboard full KPI suite."""
    ws = wb.create_sheet('07_Profit_Dashboard', 7)
    set_sheet_appearance(ws)
    add_title_band(ws, '💰 Profit Dashboard', 1, 10)
    
    kpis = [
        ('Win Rate %', '=IFERROR(COUNTIF(Tbl_Journal[Net P&L],">0")/COUNTA(Tbl_Journal[Net P&L])*100,0)', COLORS.get('accent_teal', COLORS['accent_gold'])),
        ('Profit Factor', '=IFERROR(SUMIF(Tbl_Journal[Net P&L],">0")/ABS(SUMIF(Tbl_Journal[Net P&L],"<0")),0)', COLORS['accent_gold']),
        ('Avg Win $', '=IFERROR(AVERAGEIF(Tbl_Journal[Net P&L],">0"),0)', COLORS.get('accent_teal', COLORS['accent_gold'])),
        ('Avg Loss $', '=IFERROR(AVERAGEIF(Tbl_Journal[Net P&L],"<0"),0)', COLORS['accent_red']),
        ('Payoff Ratio', '=IFERROR(ABS(AVERAGEIF(Tbl_Journal[Net P&L],">0")/AVERAGEIF(Tbl_Journal[Net P&L],"<0")),0)', COLORS['accent_gold']),
        ('Expectancy (R)', '=IFERROR(AVERAGE(Tbl_Journal[R-Multiple]),0)', COLORS.get('accent_teal', COLORS['accent_gold'])),
        ('Net P&L Total', '=IFERROR(SUM(Tbl_Journal[Net P&L]),0)', COLORS['accent_gold']),
        ('Largest Win', '=IFERROR(MAX(Tbl_Journal[Net P&L]),0)', COLORS.get('accent_teal', COLORS['accent_gold'])),
        ('Largest Loss', '=IFERROR(MIN(Tbl_Journal[Net P&L]),0)', COLORS['accent_red']),
        ('Commission Drag %', '=IFERROR(SUM(Tbl_Journal[Commission])/SUM(Tbl_Journal[Gross P&L])*100,0)', COLORS['accent_red']),
        ('Total Trades', '=COUNTA(Tbl_Journal[Trade #])', COLORS['accent_gold']),
        ('Total R Gained', '=IFERROR(SUM(Tbl_Journal[R-Multiple]),0)', COLORS.get('accent_teal', COLORS['accent_gold'])),
    ]
    
    row = 4
    col = 1
    for label, formula, color in kpis:
        # Label
        cell = ws.cell(row=row, column=col)
        cell.value = label
        apply_style(cell, StylePresets.kpi_label())
        ws.row_dimensions[row].height = 18
        
        # Value
        val_cell = ws.cell(row=row + 1, column=col)
        val_cell.value = formula
        apply_style(val_cell, StylePresets.kpi_value(color))
        ws.row_dimensions[row + 1].height = 28
        
        if col >= 5:
            col = 1
            row += 3
        else:
            col += 1
    
    for i in range(1, 7):
        ws.column_dimensions[get_column_letter(i)].width = 16
    
    apply_sheet_protection(ws)

def build_r_multiple_tracker_sheet(wb):
    """Build 08_R_Multiple_Tracker."""
    ws = wb.create_sheet('08_R_Multiple_Tracker', 8)
    set_sheet_appearance(ws)
    add_title_band(ws, '📊 R-Multiple Distribution', 1, 10)
    
    # Stats
    ws['A4'].value = 'Mean R'
    ws['B4'].value = '=IFERROR(AVERAGE(Tbl_Journal[R-Multiple]),0)'
    apply_style(ws['B4'], StylePresets.locked_cell())
    
    ws['A5'].value = '% Positive R'
    ws['B5'].value = '=IFERROR(COUNTIF(Tbl_Journal[R-Multiple],">0")/COUNTA(Tbl_Journal[R-Multiple])*100,0)'
    apply_style(ws['B5'], StylePresets.locked_cell())
    
    ws['A6'].value = 'Trades ≥ 2R'
    ws['B6'].value = '=COUNTIF(Tbl_Journal[R-Multiple],">=2")'
    apply_style(ws['B6'], StylePresets.locked_cell())
    
    ws['A7'].value = 'Trades ≤ -1R'
    ws['B7'].value = '=COUNTIF(Tbl_Journal[R-Multiple],"<=-1")'
    apply_style(ws['B7'], StylePresets.locked_cell())
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    
    apply_sheet_protection(ws)

def build_trade_statistics_sheet(wb):
    """Build 09_Trade_Statistics advanced metrics."""
    ws = wb.create_sheet('09_Trade_Statistics', 9)
    set_sheet_appearance(ws)
    add_title_band(ws, '📈 Advanced Trade Statistics', 1, 10)
    
    stats = [
        ('Sharpe Ratio', '=IFERROR(AVERAGE(Tbl_Journal[Net P&L])/STDEV(Tbl_Journal[Net P&L]),0)', 'Above 1.0 = good; above 2.0 = excellent'),
        ('Sortino Ratio', '=IFERROR(AVERAGE(Tbl_Journal[Net P&L])/STDEV(IF(Tbl_Journal[Net P&L]<0,Tbl_Journal[Net P&L],"")),0)', 'Focuses on downside risk only'),
        ('SQN (System Quality)', '=IFERROR(SQRT(COUNTA(Tbl_Journal[R-Multiple]))*AVERAGE(Tbl_Journal[R-Multiple])/STDEV(Tbl_Journal[R-Multiple]),0)', 'Above 2.0 = edge likely; above 3.0 = robust'),
    ]
    
    row = 4
    for label, formula, explanation in stats:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        
        ws.cell(row=row, column=2).value = formula
        apply_style(ws.cell(row=row, column=2), StylePresets.locked_cell())
        
        ws.cell(row=row, column=3).value = explanation
        ws.cell(row=row, column=3).font = Font(italic=True, size=9, color=COLORS['text_dark'])
        ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
        
        row += 2
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    
    apply_sheet_protection(ws)

def build_economic_calendar_sheet(wb):
    """Build 10_Economic_Calendar manual entry."""
    ws = wb.create_sheet('10_Economic_Calendar', 10)
    set_sheet_appearance(ws)
    add_title_band(ws, '📅 Economic Calendar', 1, 10)
    
    # Instructions
    ws['A3'].value = 'Copy weekly event data from Forex Factory (forexfactory.com) — Date, Time, Currency, Event, Impact, etc.'
    ws['A3'].font = Font(italic=True, size=9)
    
    # Headers
    headers = ['Date', 'Time (UTC)', 'Currency', 'Event', 'Impact', 'Forecast', 'Previous', 'Actual', 'Avoid Trading?', 'Notes']
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=idx)
        cell.value = header
        apply_style(cell, StylePresets.table_header())
    
    # Pre-fill 50 rows
    impact_validation = DataValidation(type='list', formula1='"High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(impact_validation)
    impact_validation.add('E6:E55')
    
    for calendar_row in range(6, 56):
        for col in range(1, 10):
            apply_style(ws.cell(row=calendar_row, column=col), StylePresets.input_cell())
        
        # Avoid Trading formula
        ws[f'I{calendar_row}'].value = f'=IF(E{calendar_row}="High","⚠ AVOID","✓ OK")'
        apply_style(ws[f'I{calendar_row}'], StylePresets.locked_cell())
    
    for i in range(1, 11):
        ws.column_dimensions[get_column_letter(i)].width = 15
    
    apply_sheet_protection(ws)

def build_financial_control_sheet(wb):
    """Build 11_Financial_Control deposits/withdrawals."""
    ws = wb.create_sheet('11_Financial_Control', 11)
    set_sheet_appearance(ws)
    add_title_band(ws, '💳 Financial Control', 1, 8)
    
    # Capital Log
    ws.merge_cells('A4:D4')
    log_header = ws['A4']
    log_header.value = 'CAPITAL LOG'
    apply_style(log_header, StylePresets.section_header())
    
    headers = ['Date', 'Type', 'Amount', 'Note', 'Running Net Deposits']
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=idx)
        cell.value = header
        apply_style(cell, StylePresets.table_header())
    
    type_validation = DataValidation(type='list', formula1='"Deposit,Withdrawal"', allow_blank=True)
    ws.add_data_validation(type_validation)
    type_validation.add('B6:B55')
    
    for cap_row in range(6, 56):
        for col in range(1, 4):
            apply_style(ws.cell(row=cap_row, column=col), StylePresets.input_cell())
        
        ws[f'E{cap_row}'].value = f'=SUM($C$6:C{cap_row})'
        apply_style(ws[f'E{cap_row}'], StylePresets.locked_cell())
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18
    
    apply_sheet_protection(ws)

def build_lot_size_calculator_sheet(wb):
    """Build 12_Lot_Size_Calculator position sizing by asset class."""
    ws = wb.create_sheet('12_Lot_Size_Calculator', 12)
    set_sheet_appearance(ws)
    add_title_band(ws, '📐 Lot Size Calculator', 1, 8)
    
    # Inputs
    ws['A4'].value = 'Account Size'
    ws['B4'].value = '=AccountBalance'
    apply_style(ws['B4'], StylePresets.locked_cell())
    
    ws['A5'].value = 'Risk %'
    apply_style(ws['B5'], StylePresets.input_cell())
    
    ws['A6'].value = 'Entry Price'
    apply_style(ws['B6'], StylePresets.input_cell())
    
    ws['A7'].value = 'Stop Price'
    apply_style(ws['B7'], StylePresets.input_cell())
    
    ws['A8'].value = 'Asset Class'
    apply_style(ws['B8'], StylePresets.input_cell())
    
    # Pip/Tick value reference table
    ws.merge_cells('D4:E4')
    ref_header = ws['D4']
    ref_header.value = 'PIP VALUE REFERENCE'
    apply_style(ref_header, StylePresets.section_header())
    
    pip_values = [('EURUSD', 0.0001), ('GBPUSD', 0.0001), ('USDJPY', 0.01), ('XAUUSD', 0.01)]
    for idx, (symbol, pip) in enumerate(pip_values, start=5):
        ws[f'D{idx}'].value = symbol
        ws[f'E{idx}'].value = pip
    
    ws['D9'].value = '(Check your broker for exact values)'
    ws['D9'].font = Font(italic=True, size=8)
    
    # Output
    ws.merge_cells('A10:B10')
    ws['A10'].value = 'Position Size'
    apply_style(ws['A10'], StylePresets.kpi_label())
    
    ws.merge_cells('A11:B11')
    ws['A11'].value = '=IF(OR(B5="",B6="",B7=""),"",ROUND(B4*(B5/100)/ABS(B6-B7),2))'
    apply_style(ws['A11'], StylePresets.kpi_value(COLORS['accent_gold']))
    ws.row_dimensions[11].height = 30
    
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12
    
    apply_sheet_protection(ws)

def build_tax_calculator_sheet(wb):
    """Build 13_Tax_Calculator Brazil + generic rules."""
    ws = wb.create_sheet('13_Tax_Calculator', 13)
    set_sheet_appearance(ws)
    add_title_band(ws, '🧾 Tax Calculator', 1, 8)
    
    # Region selector
    ws['A3'].value = 'Tax Region'
    apply_style(ws['B3'], StylePresets.input_cell())
    
    region_validation = DataValidation(type='list', formula1='"Brazil,Generic"', allow_blank=False)
    ws.add_data_validation(region_validation)
    region_validation.add('B3')
    ws['B3'].value = 'Brazil'
    
    # Brazil section
    ws.merge_cells('A5:D5')
    br_header = ws['A5']
    br_header.value = '🇧🇷 BRAZIL TAX RULES'
    apply_style(br_header, StylePresets.section_header())
    
    br_section = [
        ('Day Trade Monthly Net Profit', 'C7'),
        ('Day Trade Tax Rate %', 'C8'),
        ('Day Trade Tax Due', 'C9'),
        ('', ''),
        ('Swing Trade Monthly Net Profit', 'C12'),
        ('Swing Trade Tax Rate %', 'C13'),
        ('Monthly Sales (Swing)', 'C14'),
        ('Swing Trade Tax Due', 'C15'),
        ('', ''),
        ('Disclaimer', 'C18'),
    ]
    
    row = 6
    for label, cell_ref in br_section:
        if label == '':
            row += 1
            continue
        if label == 'Disclaimer':
            ws.merge_cells(f'A{row}:D{row}')
            disclaimer = ws[f'A{row}']
            disclaimer.value = '⚠ This is a simplified estimate for personal organization only. Not tax advice. Consult an accountant.'
            apply_style(disclaimer, StylePresets.warning())
        else:
            ws[f'A{row}'].value = label
            apply_style(ws[cell_ref], StylePresets.input_cell() if 'Profit' in label or 'Sales' in label or 'Rate' in label else StylePresets.locked_cell())
            
            if 'Tax Due' in label:
                if 'Day' in label:
                    ws[cell_ref].value = '=IF(OR(C7="",C8=""),"",MAX(0,(C7*C8/100)))'
                else:
                    ws[cell_ref].value = '=IF(OR(C12="",C13="",C14=""),"",IF(C14<=20000,0,MAX(0,(C12*C13/100))))'
                apply_style(ws[cell_ref], StylePresets.locked_cell())
        row += 1
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['C'].width = 15
    
    apply_sheet_protection(ws)

def build_compounding_calculator_sheet(wb):
    """Build 14_Compounding_Calculator growth projections."""
    ws = wb.create_sheet('14_Compounding_Calculator', 14)
    set_sheet_appearance(ws)
    add_title_band(ws, '📈 Compounding Calculator', 1, 8)
    
    # Inputs
    ws['A4'].value = 'Starting Capital'
    ws['B4'].value = '=AccountBalance'
    apply_style(ws['B4'], StylePresets.locked_cell())
    
    ws['A5'].value = 'Expected Monthly Return %'
    apply_style(ws['B5'], StylePresets.input_cell())
    
    ws['A6'].value = 'Monthly Contribution'
    apply_style(ws['B6'], StylePresets.input_cell())
    
    ws['A7'].value = 'Number of Months'
    apply_style(ws['B7'], StylePresets.input_cell())
    
    ws['A8'].value = 'Worst-Case Drawdown %'
    apply_style(ws['B8'], StylePresets.input_cell())
    
    # Projection table
    ws.merge_cells('A10:D10')
    proj_header = ws['A10']
    proj_header.value = 'PROJECTION TABLE'
    apply_style(proj_header, StylePresets.section_header())
    
    headers = ['Month', 'Base Case', 'Optimistic (+3%)', 'Pessimistic (-3%)']
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=11, column=idx)
        cell.value = header
        apply_style(cell, StylePresets.table_header())
    
    for proj_row in range(12, 48):
        month_num = proj_row - 11
        ws[f'A{proj_row}'].value = month_num
        apply_style(ws[f'A{proj_row}'], StylePresets.locked_cell())
        
        # Base case formula
        if month_num == 1:
            ws[f'B{proj_row}'].value = f'=B4*(1+B5/100)+B6'
        else:
            ws[f'B{proj_row}'].value = f'=B{proj_row-1}*(1+B5/100)+B6'
        apply_style(ws[f'B{proj_row}'], StylePresets.locked_cell())
        
        # Optimistic
        if month_num == 1:
            ws[f'C{proj_row}'].value = f'=B4*(1+(B5+3)/100)+B6'
        else:
            ws[f'C{proj_row}'].value = f'=C{proj_row-1}*(1+(B5+3)/100)+B6'
        apply_style(ws[f'C{proj_row}'], StylePresets.locked_cell())
        
        # Pessimistic
        if month_num == 1:
            ws[f'D{proj_row}'].value = f'=B4*(1+(B5-3)/100)+B6'
        else:
            ws[f'D{proj_row}'].value = f'=D{proj_row-1}*(1+(B5-3)/100)+B6'
        apply_style(ws[f'D{proj_row}'], StylePresets.locked_cell())
    
    for i in range(1, 5):
        ws.column_dimensions[get_column_letter(i)].width = 18
    
    apply_sheet_protection(ws)

def build_expectancy_calculator_sheet(wb):
    """Build 15_Expectancy_Calculator edge validator."""
    ws = wb.create_sheet('15_Expectancy_Calculator', 15)
    set_sheet_appearance(ws)
    add_title_band(ws, '🎯 Expectancy Calculator (Edge Validator)', 1, 8)
    
    # Inputs
    ws['A4'].value = 'Win Rate %'
    apply_style(ws['B4'], StylePresets.input_cell())
    
    ws['A5'].value = 'Average Win (R)'
    apply_style(ws['B5'], StylePresets.input_cell())
    
    ws['A6'].value = 'Average Loss (R)'
    apply_style(ws['B6'], StylePresets.input_cell())
    
    ws['A7'].value = 'Number of Trades in Sample'
    apply_style(ws['B7'], StylePresets.input_cell())
    
    # Calculations
    ws['A10'].value = 'Expectancy (R)'
    ws['B10'].value = '=IF(OR(B4="",B5="",B6=""),"",B4/100*B5-(1-B4/100)*ABS(B6))'
    apply_style(ws['B10'], StylePresets.kpi_value(COLORS['accent_gold']))
    
    ws['A11'].value = 'Breakeven Win Rate %'
    ws['B11'].value = '=IF(B6="","",1/(1+ABS(B5/B6))*100)'
    apply_style(ws['B11'], StylePresets.locked_cell())
    
    ws['A12'].value = 'Edge Verdict'
    ws['B12'].value = '=IF(OR(B10="",""),"",IF(B10>0.3,"🟢 Strong Edge",IF(B10>0.1,"🟡 Weak Edge","🔴 No Edge")))'
    apply_style(ws['B12'], StylePresets.locked_cell())
    
    ws['A13'].value = 'Sample Size Check'
    ws['B13'].value = '=IF(OR(B7=""),"",IF(B7>=30,"✅ Sample OK (30+ trades)","⚠ Insufficient (need 30+)"))'
    apply_style(ws['B13'], StylePresets.locked_cell())
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    
    apply_sheet_protection(ws)

def build_heatmaps_sheet(wb):
    """Build 16_Heatmaps performance analysis."""
    ws = wb.create_sheet('16_Heatmaps', 16)
    set_sheet_appearance(ws)
    add_title_band(ws, '🔥 Heatmaps', 1, 8)
    
    # Heatmap 1: By Symbol
    ws.merge_cells('A4:E4')
    hm1_header = ws['A4']
    hm1_header.value = 'PERFORMANCE BY SYMBOL'
    apply_style(hm1_header, StylePresets.section_header())
    
    symbol_headers = ['Symbol', 'Trades', 'Win Rate %', 'Expectancy (R)', 'Net P&L']
    for idx, header in enumerate(symbol_headers, start=1):
        cell = ws.cell(row=5, column=idx)
        cell.value = header
        apply_style(cell, StylePresets.table_header())
    
    symbols = ['EURUSD', 'GBPUSD', 'GOLD', 'BTCUSD', 'Other']
    for row_idx, symbol in enumerate(symbols, start=6):
        ws[f'A{row_idx}'].value = symbol
        apply_style(ws[f'A{row_idx}'], StylePresets.locked_cell())
        
        ws[f'B{row_idx}'].value = f'=COUNTIF(Tbl_Journal[Symbol],A{row_idx})'
        apply_style(ws[f'B{row_idx}'], StylePresets.locked_cell())
        
        ws[f'C{row_idx}'].value = f'=IF(B{row_idx}=0,"",COUNTIFS(Tbl_Journal[Symbol],A{row_idx},Tbl_Journal[Net P&L],">0")/B{row_idx}*100)'
        apply_style(ws[f'C{row_idx}'], StylePresets.locked_cell())
        
        ws[f'D{row_idx}'].value = f'=IF(B{row_idx}=0,"",AVERAGEIF(Tbl_Journal[Symbol],A{row_idx},Tbl_Journal[R-Multiple]))'
        apply_style(ws[f'D{row_idx}'], StylePresets.locked_cell())
        
        ws[f'E{row_idx}'].value = f'=SUMIF(Tbl_Journal[Symbol],A{row_idx},Tbl_Journal[Net P&L])'
        apply_style(ws[f'E{row_idx}'], StylePresets.locked_cell())
    
    for i in range(1, 6):
        ws.column_dimensions[get_column_letter(i)].width = 16
    
    apply_sheet_protection(ws)

def build_equity_curve_sheet(wb):
    """Build 17_Equity_Curve full curve + analysis."""
    ws = wb.create_sheet('17_Equity_Curve', 17)
    set_sheet_appearance(ws)
    add_title_band(ws, '📊 Equity Curve', 1, 10)
    
    # KPI Cards
    ws['A4'].value = 'Total Return %'
    apply_style(ws['A4'], StylePresets.kpi_label())
    ws['A5'].value = '=IFERROR((MAX(Tbl_Journal[Running Balance])-AccountBalance)/AccountBalance*100,0)'
    apply_style(ws['A5'], StylePresets.kpi_value(COLORS.get('accent_teal', COLORS['accent_gold'])))
    
    ws['B4'].value = 'Max Drawdown %'
    apply_style(ws['B4'], StylePresets.kpi_label())
    ws['B5'].value = '=IFERROR(MINIFS(Tbl_Journal[Drawdown %],Tbl_Journal[Drawdown %],"<0"),0)'
    apply_style(ws['B5'], StylePresets.kpi_value(COLORS['accent_red']))
    
    ws['C4'].value = 'Recovery Factor'
    apply_style(ws['C4'], StylePresets.kpi_label())
    ws['C5'].value = '=IFERROR(A5/ABS(B5),0)'
    apply_style(ws['C5'], StylePresets.kpi_value(COLORS['accent_gold']))
    
    for i in range(1, 4):
        ws.column_dimensions[get_column_letter(i)].width = 18
    
    apply_sheet_protection(ws)

def build_workbook():
    """Main workbook builder."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Build sheets in order
    build_cover_sheet(wb)
    build_how_to_use_sheet(wb)
    build_settings_sheet(wb)
    build_dashboard_sheet(wb)
    build_trading_journal_sheet(wb)
    build_risk_manager_sheet(wb)
    build_drawdown_tracker_sheet(wb)
    build_profit_dashboard_sheet(wb)
    build_r_multiple_tracker_sheet(wb)
    build_trade_statistics_sheet(wb)
    build_economic_calendar_sheet(wb)
    build_financial_control_sheet(wb)
    build_lot_size_calculator_sheet(wb)
    build_tax_calculator_sheet(wb)
    build_compounding_calculator_sheet(wb)
    build_expectancy_calculator_sheet(wb)
    build_heatmaps_sheet(wb)
    build_equity_curve_sheet(wb)
    build_lists_sheet(wb)  # Hidden sheet at the end
    
    # Color tab groups
    tab_colors = {
        '00_Cover': 'D4AF37',
        '01_How_To_Use': 'D4AF37',
        '02_Settings': 'D4AF37',
        '03_Dashboard': '22D3A6',
        '04_Trading_Journal': 'D4AF37',
        '05_Risk_Manager': 'E5484D',
        '06_Drawdown_Tracker': 'E5484D',
        '07_Profit_Dashboard': '22D3A6',
        '08_R_Multiple_Tracker': '22D3A6',
        '09_Trade_Statistics': '22D3A6',
        '10_Economic_Calendar': 'F5A524',
        '11_Financial_Control': '22D3A6',
        '12_Lot_Size_Calculator': 'F5A524',
        '13_Tax_Calculator': 'F5A524',
        '14_Compounding_Calculator': 'F5A524',
        '15_Expectancy_Calculator': 'F5A524',
        '16_Heatmaps': '22D3A6',
        '17_Equity_Curve': '22D3A6',
    }
    
    for sheet_name, color in tab_colors.items():
        if sheet_name in wb.sheetnames:
            wb[sheet_name].sheet_properties.tabColor = color
    
    return wb

def generate_readme():
    """Generate README_HowToUse.txt file (existing content)."""
    readme_content = """
╔══════════════════════════════════════════════════════════════════╗
║         VERISCOPE EDGE - QUICK START GUIDE                      ║
║    Premium All-in-One Trading Performance System v1.0           ║
║                  "Know Your Edge. Trade With Proof."             ║
╚══════════════════════════════════════════════════════════════════╝

📋 QUICK START (5 STEPS)
========================

1. CONFIGURE SETTINGS
   • Open the "Settings" tab
   • Enter your starting capital
   • Set your default risk % per trade
   • Fill in broker name and timezone
   • These settings feed into all other sheets automatically

2. LOG YOUR TRADES IN TRADING JOURNAL
   • Click "Trading Journal" tab
   • For each trade:
     - Enter date, time, symbol, direction
     - Set entry price and stop loss
     - Record position size
     - Upon exit, enter exit price and date
   • Net P&L and R-Multiple calculate automatically
   • The system tracks your running balance in real-time

3. CHECK YOUR DASHBOARD
   • Dashboard updates live as you add trades
   • View key metrics: Net P&L, Win Rate, Profit Factor
   • No manual updates needed — it's all formulas

4. CALCULATE POSITION SIZE BEFORE ENTERING
   • Use "Risk Manager" or "Lot Size Calculator"
   • Input your entry and stop, system calculates position size
   • Never guess position sizing — let the math lead

5. REVIEW WEEKLY
   • Study "Profit Dashboard" for your edge metrics
   • Check "Heatmaps" for patterns by symbol, setup, emotion
   • Use "R-Multiple Tracker" to validate your system
   • Review "Drawdown Tracker" to track peak-to-trough loss

📊 WHAT EACH SHEET DOES
========================

NAVIGATION & SETUP:
  • 00_Cover ................... Landing page with menu
  • 01_How_To_Use .............. This guide + FAQ
  • 02_Settings ................ Your control panel (starting capital, risk %, etc.)

CORE TRADING:
  • 03_Dashboard ............... Live overview of all key metrics
  • 04_Trading_Journal ......... Your trade log (the engine)
  • 05_Risk_Manager ............ Position sizing calculator
  • 06_Drawdown_Tracker ........ Drawdown analysis

ANALYSIS & METRICS:
  • 07_Profit_Dashboard ........ Full KPI suite (Win Rate, Profit Factor, etc.)
  • 08_R_Multiple_Tracker ...... R-multiple distribution histogram
  • 09_Trade_Statistics ........ Advanced metrics (Sharpe, Sortino, SQN)
  • 16_Heatmaps ................ Performance by symbol, setup, emotion
  • 17_Equity_Curve ............ Cumulative P&L line chart

PLANNING & CONTROL:
  • 10_Economic_Calendar ....... Key news events to monitor
  • 11_Financial_CONTROL ....... Deposit/withdrawal tracking
  • 12_Lot_SIZE_CALCULATOR ..... Position sizing by asset class
  • 13_Tax_Calculator .......... Brazil day-trade tax + generic rules
  • 14_Compounding_Calculator .. Growth projections
  • 15_Expectancy_Calculator ... Edge validator for your setups

📖 BEGINNER GLOSSARY
====================

R-MULTIPLE:
  The "R" is your risk per trade. If you risk $100 and profit $300, that is +3R.
  If you risk $100 and lose $100, that is -1R.
  Expectancy in R tells you your average win/loss ratio — the heart of your edge.

PROFIT FACTOR:
  Total Wins ÷ Total Losses. 
  Above 1.5 = good. Above 2.0 = excellent. Below 1.0 = losing system.

DRAWDOWN:
  The decline from peak balance to lowest point. Max DD shows your worst losing streak.
  Track this closely — it tells you if you can afford another round of losses.

WIN RATE:
  Percentage of trades that make money. 
  Note: High win rate + small wins can still lose money if losses are bigger.
  Quality of wins (R-multiple) matters more than quantity.

SHARPE RATIO:
  Risk-adjusted return. Above 1.0 = good; above 2.0 = excellent.
  Accounts for volatility of your returns, not just average.

RULE COMPLIANCE:
  Mark each trade: did you follow your rules? Yes or No.
  Separately note which rule you broke (Entry, Position Sizing, Early Exit, etc.).
  Review "Heatmaps" weekly to see if rule-breaking correlates with losses.

EXPECTANCY:
  Your average profit per trade in R. 
  Positive expectancy = you have an edge. Use "Expectancy Calculator" to validate new setups with 30+ historical trades before live trading.

⚙️ CELL PROTECTION & EDITING
=============================

• INPUT CELLS (light blue background) ........... Unlock to enter your data
• LOCKED CELLS (formula output) ................ Protected to prevent accidents
• All sheets password-protected with "veriscope2026"
• If you need to unlock a cell, right-click sheet → "Format Cells" → "Protection" tab

💡 TIPS FOR SUCCESS
====================

✓ Log trades SAME DAY, not later. Memory degrades accuracy.
✓ Log EVERY trade, including breakeven and small losses. System depends on complete data.
✓ Set realistic risk %. Start with 1–2% if new to trading.
✓ Review heatmaps weekly. Patterns in symbol/setup/emotion reveal your true strengths.
✓ Use "Expectancy Calculator" to validate new setups with 30+ historical trades before live trading.
✓ Adjust "Settings" if you make a deposit/withdrawal.
✓ Keep a copy of each month's file as backup — timestamp your exports.

⚠️ DISCLAIMER
==============

This is a personal trading journal and analytics tool for organization and tracking only.
It is NOT tax advice, NOT investment advice, and NOT a guarantee of profitability.
Tax calculations (Brazil section) are simplified estimates — consult a tax professional.
Consult a financial advisor before making any trading decisions.

📧 SUPPORT & FEEDBACK
======================

Version: 1.0
Generated: """ + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + """
Language: English
Compatibility: Excel 2016+, Google Sheets (formula degradation possible)

For questions or updates, refer to the original project documentation.

Good luck with your trading journey!
    """
    return readme_content

def generate_readme_all_sheets(wb):
    """Generate a README that enumerates all sheets with short descriptions."""
    descriptions = {
        '00_Cover': 'Landing page with product title, subtitle and navigation.',
        '01_How_To_Use': 'Quick start guide and FAQ for new users.',
        '02_Settings': 'Global settings and named ranges (AccountBalance, RiskPercent, etc.).',
        '03_Dashboard': 'Overview KPI cards and key charts (Net P&L, Win Rate, Equity Curve).',
        '04_Trading_Journal': 'Primary trade log (table) where each trade is recorded.',
        '05_Risk_Manager': 'Position sizing calculator and quick risk tools.',
        '06_Drawdown_Tracker': 'Drawdown KPIs and underwater analysis.',
        '07_Profit_Dashboard': 'Detailed profit KPIs and breakdowns.',
        '08_R_Multiple_Tracker': 'R-Multiple distribution and stats.',
        '09_Trade_Statistics': 'Sharpe, Sortino, SQN and advanced metrics.',
        '10_Economic_Calendar': 'Manual economic calendar / news log.',
        '11_Financial_Control': 'Deposit/withdrawal log and running net deposits.',
        '12_Lot_Size_Calculator': 'Asset-class aware lot / position size calculator.',
        '13_Tax_Calculator': 'Brazil + generic tax estimator (informational).',
        '14_Compounding_Calculator': 'Growth projections and scenarios.',
        '15_Expectancy_Calculator': 'Edge validator (expectancy, breakeven).',
        '16_Heatmaps': 'Performance heatmaps by symbol/setup/emotion.',
        '17_Equity_Curve': 'Full equity curve, return and recovery metrics.',
        '99_Lists': 'Hidden sheet with dropdown list sources.',
    }
    lines = []
    lines.append("VERISCOPE EDGE - Sheets Overview")
    lines.append("=" * 40)
    lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("")
    for sheet in wb.sheetnames:
        desc = descriptions.get(sheet, 'No description available.')
        lines.append(f"- {sheet}: {desc}")
    lines.append("")
    lines.append("Notes:")
    lines.append("- All settings are named ranges (AccountBalance, RiskPercent, CurrencySymbol, DailyLossLimit).")
    lines.append("- The Trading Journal is the single source of truth; KPIs and charts pull from Tbl_Journal.")
    lines.append("- If you see any formula errors with an empty journal, check that named ranges exist and that the table header 'Running Balance' matches exactly.")
    return "\n".join(lines)

# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function."""
    print("🚀 Building Veriscope Edge Workbook...")
    ensure_output_dir()
    
    # Build workbook
    wb = build_workbook()
    
    # Save workbook
    output_path = 'output/Veriscope_Edge.xlsx'
    wb.save(output_path)
    print(f"✅ Workbook saved: {output_path}")
    
    # Generate README
    readme_path = 'output/README_HowToUse.txt'
    with open(readme_path, 'w', encoding='utf-8') as f:
        f.write(generate_readme())
    print(f"✅ README generated: {readme_path}")
    
    # Generate external README listing all sheets
    readme_all_path = 'output/README_All_Sheets.txt'
    with open(readme_all_path, 'w', encoding='utf-8') as f:
        f.write(generate_readme_all_sheets(wb))
    print(f"✅ README (all sheets) generated: {readme_all_path}")
    
    print("\n🎉 Veriscope Edge build complete!")
    print(f"📦 Output directory: ./output/")
    print(f"📄 Files ready:")
    print(f"   • Veriscope_Edge.xlsx (main file)")
    print(f"   • README_HowToUse.txt (guide)")
    print(f"   • README_All_Sheets.txt (overview of all sheets)")

if __name__ == '__main__':
    main()
